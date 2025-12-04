"""
Utility functions for importing users from Excel files.
Supports Vietnamese user profile fields and bulk user creation.
"""

import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Tuple

from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from common.djangoapps.student.models import UserProfile

logger = logging.getLogger(__name__)

# Column mappings: Excel header -> UserProfile field/meta key
EXCEL_COLUMN_MAPPINGS = {
    'Họ và tên': 'name',
    'Ngày sinh': 'ngay_sinh',
    'Giới tính': 'gender',
    'Điện thoại': 'phone_number',
    'Email': 'email',
    'CCCD': 'cccd',
    'Ngày cấp CCCD': 'ngay_cap_cccd',
    'Đơn vị công tác': 'don_vi_cong_tac',
    'Quê quán': 'que_quan',
    'Dân tộc': 'dan_toc',
    'Ghi chú': 'ghi_chu',
    'Ảnh avatar': 'avatar_url',
    'Mật khẩu': 'password',
    'Vị trí việc làm': 'vi_tri_viec_lam',
    'Chức vụ': 'chuc_vu',
    'Người nhận bằng': 'nguoi_nhan_bang',
    'Số chứng chỉ': 'so_chung_chi',
    'Tên khóa học': 'ten_khoa_hoc',
    'Thời gian học': 'thoi_gian_hoc',
    'Năm tốt nghiệp': 'nam_tot_nghiep',
    'Số tiết quy đổi': 'so_tiet_quy_doi',
    'Loại hình đào tạo': 'loai_hinh_dao_tao',
    'Nơi sinh': 'noi_sinh',
    'Địa chỉ': 'dia_chi',
    'Số năm công tác': 'so_nam_cong_tac',
    'Vai trò người dùng hệ thống': 'user_role',  # Quản trị cơ quan | Giảng viên | Học viên
}

# Fields that should be stored in UserProfile model directly (not in meta)
DIRECT_PROFILE_FIELDS = {'name', 'phone_number', 'gender'}

# Fields that go to User model
USER_MODEL_FIELDS = {'email', 'password'}

# Special fields that require custom handling
SPECIAL_FIELDS = {'user_role'}  # Requires role assignment

# Fields that should be stored in UserProfile.meta as JSON
META_FIELDS = set(EXCEL_COLUMN_MAPPINGS.values()) - DIRECT_PROFILE_FIELDS - USER_MODEL_FIELDS - SPECIAL_FIELDS


def generate_excel_template() -> bytes:
    """
    Generate an Excel template file with required columns.
    
    Returns:
        bytes: Excel file content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách người dùng"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    headers = list(EXCEL_COLUMN_MAPPINGS.keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add sample data row with instructions
    sample_row = {
        'Họ và tên': 'Nguyễn Văn A',
        'Ngày sinh': '01/01/1990',
        'Giới tính': 'm',
        'Điện thoại': '0123456789',
        'Email': 'nguyenvana@example.com',
        'CCCD': '001234567890',
        'Ngày cấp CCCD': '01/01/2020',
        'Đơn vị công tác': 'Tên tổ chức phải khớp chính xác với tổ chức đã tạo trong hệ thống',
        'Quê quán': 'Hà Nội',
        'Dân tộc': 'Kinh',
        'Ghi chú': '',
        'Ảnh avatar': 'https://example.com/avatar.jpg',
        'Mật khẩu': 'password123',
        'Vị trí việc làm': 'Giáo viên',
        'Chức vụ': 'Trưởng phòng',
        'Người nhận bằng': 'Nguyễn Văn A',
        'Số chứng chỉ': 'CC-001',
        'Tên khóa học': 'Khóa học A',
        'Thời gian học': '6 tháng',
        'Năm tốt nghiệp': '2020',
        'Số tiết quy đổi': '120',
        'Loại hình đào tạo': 'Trực tiếp',
        'Nơi sinh': 'Hà Nội',
        'Địa chỉ': '123 Đường ABC, Hà Nội',
        'Số năm công tác': '5',
        'Vai trò người dùng hệ thống': 'Quản trị cơ quan hoặc Giảng viên hoặc Học viên hoặc Công chức',
    }
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = sample_row.get(header, '')
    
    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_num)].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_excel_file(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parse Excel file and extract user data.
    
    Args:
        file_content: Excel file content as bytes
        
    Returns:
        List of dictionaries containing user data
        
    Raises:
        ValueError: If file format is invalid
    """
    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        # Read headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Validate headers
        expected_headers = set(EXCEL_COLUMN_MAPPINGS.keys())
        provided_headers = set(headers)
        missing_headers = expected_headers - provided_headers
        
        if missing_headers:
            raise ValueError(f"Missing required columns: {', '.join(missing_headers)}")
        
        # Parse data rows
        users_data = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            user_data = {}
            for col_num, value in enumerate(row):
                if col_num < len(headers):
                    header = headers[col_num]
                    field_name = EXCEL_COLUMN_MAPPINGS.get(header)
                    if field_name:
                        # Convert cell value to string, handling None
                        user_data[field_name] = str(value).strip() if value is not None else ''
            
            # Add row number for error reporting
            user_data['_row_number'] = row_num
            users_data.append(user_data)
        
        return users_data
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise ValueError(f"Không thể đọc file Excel: {str(e)}")


def validate_user_data(user_data: Dict[str, Any]) -> Tuple[bool, List[str]]:
    """
    Validate user data from Excel row.
    
    Args:
        user_data: Dictionary containing user data
        
    Returns:
        Tuple of (is_valid, error_messages)
    """
    errors = []
    row_num = user_data.get('_row_number', '?')
    
    # Validate required fields
    required_fields = ['name', 'email', 'password']
    for field in required_fields:
        if not user_data.get(field):
            errors.append(f"Dòng {row_num}: Thiếu trường bắt buộc '{field}'")
    
    # Validate email format
    email = user_data.get('email', '').strip()
    if email:
        try:
            validate_email(email)
        except ValidationError:
            errors.append(f"Dòng {row_num}: Email không hợp lệ '{email}'")
        
        # Check if email already exists
        if User.objects.filter(email=email).exists():
            errors.append(f"Dòng {row_num}: Email '{email}' đã tồn tại trong hệ thống")
    
    # Validate password strength
    password = user_data.get('password', '').strip()
    if password and len(password) < 6:
        errors.append(f"Dòng {row_num}: Mật khẩu phải có ít nhất 6 ký tự")
    
    # Validate gender
    gender = user_data.get('gender', '').strip().lower()
    if gender and gender not in ['m', 'f', 'o', '']:
        errors.append(f"Dòng {row_num}: Giới tính không hợp lệ '{gender}' (chỉ chấp nhận: m, f, o)")
    
    # Validate phone number format (basic)
    phone = user_data.get('phone_number', '').strip()
    if phone and not phone.replace('+', '').replace(' ', '').isdigit():
        errors.append(f"Dòng {row_num}: Số điện thoại không hợp lệ '{phone}'")
    
    return len(errors) == 0, errors


def create_user_from_data(user_data: Dict[str, Any], created_by: User, force_org: ChalixOrganization = None) -> Tuple[User, List[str]]:
    """
    Create a user account and profile from validated data.
    
    Args:
        user_data: Validated user data dictionary
        created_by: User who is creating this account
        force_org: Optional organization to force for all created users (used by org admins)
        
    Returns:
        Tuple of (created_user, warning_messages)
    """
    from cms.djangoapps.contentstore.models import ChalixUserRole
    
    warnings = []
    
    # Extract fields for User model
    email = user_data.get('email', '').strip()
    password = user_data.get('password', '').strip()
    name = user_data.get('name', '').strip()
    user_role = user_data.get('user_role', '').strip()  # Quản trị cơ quan | Giảng viên | Học viên
    don_vi_cong_tac = user_data.get('don_vi_cong_tac', '').strip()  # Organization name
    
    # Generate username from email
    username = email.split('@')[0]
    
    # Ensure username is unique
    base_username = username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}{counter}"
        counter += 1
        warnings.append(f"Username '{base_username}' đã tồn tại, đã tạo username mới: '{username}'")
    
    try:
        # Create User
        user = User.objects.create(
            username=username,
            email=email,
            password=make_password(password)
        )
        
        # Get or create UserProfile
        try:
            profile = UserProfile.objects.get(user=user)
        except UserProfile.DoesNotExist:
            profile = UserProfile(user=user)
        
        # Set direct profile fields
        profile.name = name
        
        if user_data.get('phone_number'):
            profile.phone_number = user_data['phone_number'].strip()
        
        if user_data.get('gender'):
            gender = user_data['gender'].strip().lower()
            if gender in ['m', 'f', 'o']:
                profile.gender = gender
        
        # Build meta data dictionary for additional fields
        meta_data = profile.get_meta()
        
        for field_name, field_value in user_data.items():
            if field_name in META_FIELDS and field_value:
                meta_data[field_name] = str(field_value).strip()
        
        # Save meta data
        profile.set_meta(meta_data)
        profile.save()
        
        # Assign user role if provided
        # If force_org is set (org admin uploading), use that instead of parsing from file
        if user_role:
            # Map Vietnamese role names to role codes
            role_mapping = {
                'Quản trị cơ quan': 'co_quan',
                'Giảng viên': 'giang_vien',
                'Học viên': 'cong_chuc',  # Changed from 'hoc_vien' to match model
                'Công chức': 'cong_chuc',
                'Viên chức': 'cong_chuc',
            }
            
            role_code = role_mapping.get(user_role)
            if role_code:
                try:
                    from cms.djangoapps.contentstore.models import ChalixOrganization
                    
                    organization = None
                    
                    # If force_org is provided (org admin), use it and ignore don_vi_cong_tac
                    if force_org:
                        organization = force_org
                        logger.info(f"Using force_org '{organization.display_name}' for user {username}")
                    elif don_vi_cong_tac:
                        # Find organization by name (try exact match first, then case-insensitive)
                        organization = ChalixOrganization.objects.filter(name=don_vi_cong_tac).first()
                        if not organization:
                            # Try case-insensitive match
                            organization = ChalixOrganization.objects.filter(name__iexact=don_vi_cong_tac).first()
                        
                        if not organization:
                            # Try display_name match
                            organization = ChalixOrganization.objects.filter(display_name__iexact=don_vi_cong_tac).first()
                        
                        if not organization:
                            warnings.append(f"Không tìm thấy tổ chức '{don_vi_cong_tac}' trong hệ thống. Vui lòng tạo tổ chức trước.")
                            logger.warning(f"Organization '{don_vi_cong_tac}' not found for user {username}")
                    
                    if organization:
                        # Check if user already has a role for this organization
                        existing_role = ChalixUserRole.objects.filter(
                            user=user,
                            organization=organization
                        ).first()
                        
                        if existing_role:
                            # Update existing role
                            existing_role.role = role_code
                            existing_role.is_active = True
                            existing_role.created_by = created_by
                            existing_role.save()
                            warnings.append(f"Cập nhật vai trò '{user_role}' cho người dùng {username} tại '{organization.display_name}'")
                        else:
                            # Create new role
                            ChalixUserRole.objects.create(
                                user=user,
                                role=role_code,
                                organization=organization,
                                is_active=True,
                                created_by=created_by
                            )
                            logger.info(f"Assigned role '{role_code}' to user {username} at organization '{organization.display_name}'")
                except Exception as e:
                    warnings.append(f"Không thể gán vai trò '{user_role}': {str(e)}")
                    logger.error(f"Failed to assign role to user {username}: {str(e)}", exc_info=True)
            else:
                warnings.append(f"Vai trò '{user_role}' không hợp lệ. Chỉ chấp nhận: Quản trị cơ quan, Giảng viên, Học viên, Công chức, Viên chức")
        
        logger.info(f"Created user: {username} ({email}) by {created_by.username}")
        
        return user, warnings
        
    except Exception as e:
        logger.error(f"Error creating user {email}: {str(e)}")
        raise


def import_users_from_excel(
    file_content: bytes,
    created_by: User,
    force_org: ChalixOrganization = None
) -> Dict[str, Any]:
    """
    Import multiple users from an Excel file.
    
    Args:
        file_content: Excel file content as bytes
        created_by: User who is importing
        force_org: Optional organization to force for all created users (used by org admins)
        
    Returns:
        Dictionary with import results
    """
    result = {
        'success': False,
        'total_rows': 0,
        'successful_imports': 0,
        'failed_imports': 0,
        'errors': [],
        'warnings': [],
        'created_users': []
    }
    
    try:
        # Parse Excel file
        users_data = parse_excel_file(file_content)
        result['total_rows'] = len(users_data)
        
        if not users_data:
            result['errors'].append("File Excel không chứa dữ liệu người dùng")
            return result
        
        # Validate all users first
        validation_errors = []
        for user_data in users_data:
            is_valid, errors = validate_user_data(user_data)
            if not is_valid:
                validation_errors.extend(errors)
        
        if validation_errors:
            result['errors'] = validation_errors
            return result
        
        # Create users in transaction
        with transaction.atomic():
            for user_data in users_data:
                try:
                    user, warnings = create_user_from_data(user_data, created_by, force_org)
                    result['successful_imports'] += 1
                    result['created_users'].append({
                        'username': user.username,
                        'email': user.email,
                        'name': user.profile.name
                    })
                    if warnings:
                        result['warnings'].extend(warnings)
                        
                except Exception as e:
                    result['failed_imports'] += 1
                    row_num = user_data.get('_row_number', '?')
                    result['errors'].append(f"Dòng {row_num}: Lỗi tạo tài khoản - {str(e)}")
        
        result['success'] = result['successful_imports'] > 0
        
    except ValueError as e:
        result['errors'].append(str(e))
    except Exception as e:
        logger.error(f"Unexpected error during import: {str(e)}")
        result['errors'].append(f"Lỗi không xác định: {str(e)}")
    
    return result
